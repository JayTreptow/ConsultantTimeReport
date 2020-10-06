

import csv
from datetime import datetime, date
import sys, getopt
from openpyxl import load_workbook
from openpyxl.styles.alignment import Alignment
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.dimensions import ColumnDimension

usage = "PO_Budget_Analysis.py -o <output file>\n\t-h : help \n\t-p : Project Future \n\t-c <calendar file> \n\t-s <SpringAhead file> \n\t-d <Deltek File> \n\t-e <Employee File>"
calendarFile = ""
springAheadFile = ""
deltekFile = ""
employeeFile = ""
outputFile = ""
nSpringAheadLastWeekNum = 1
firstProjectedWeek = 1
doProjection = 0

def getCmdlineArgs(argv):
  global calendarFile
  global springAheadFile
  global deltekFile
  global employeeFile
  global outputFile
  global doProjection

  try:
      opts, args = getopt.getopt(argv,"hpc:d:e:o:s:")
  except getopt.GetoptError:
      print(usage)

  for opt, arg in opts:
    if opt == '-h':
      return 1
    elif opt in("-c"):
      calendarFile = arg 
    elif opt in("-d"):
      deltekFile = arg 
    elif opt in("-s"):
      springAheadFile = arg 
    elif opt in("-e"):
      employeeFile = arg 
    elif opt in("-o"):
      outputFile = arg
    elif opt in("-p"):
      something = arg
      doProjection = 1

  if len(outputFile) <= 0:
    return 2

  return 0

def readCalendarFile():
  inData = {}  
  if len(calendarFile) <= 0:
    for w in range(1,54):
      inData[w] = float(40)
  else:
    with open(calendarFile, newline='') as fileIn:
      inDict = csv.DictReader(fileIn, delimiter=',', quotechar='"')
      for row in inDict:
        weeks = list(row.items())
        for w in range(0,53):
          inData[w+1] = weeks[w][1]

  return(inData)

def readSpringAheadFile():
  inData = {}  
  if len(springAheadFile) <= 0:
    return(inData)
  with open(springAheadFile, newline='') as fileIn:
    inDict = csv.DictReader(fileIn, delimiter=',', quotechar='"')
    for row in inDict:
      usr = row['User']
      date = datetime.strptime(row['Date'],'%m/%d/%Y')
      weekNum = date.isocalendar()[1]
      if usr in inData:
        dayHours = float(row['Hours'])
        if weekNum in inData[usr]:
          inData[usr][weekNum] = float(inData[usr][weekNum]) + dayHours
        else:
          inData[usr][weekNum] = dayHours
      else:
        stats = {}
        stats['SpringAhead Rate'] = float(row['Bill Rate'].split("$")[1])
        stats['Avg Run'] = float(40)
        stats["Hours"] = ""
        stats["Dollars"] = ""
        stats[weekNum] = float(row['Hours'])
        inData[usr] = stats

  return(inData)

def readDeltekExcelFile():
  inData = {}  
  if len(deltekFile) <= 0:
    return(inData)
  excelIn = load_workbook(deltekFile)
  ws = excelIn.worksheets[0]

  dateCol = headerRow = nameCol = rateCol = hourCol = 0
  for row in ws.iter_rows(1,10,1,30):
    for rowCell in row:
      rowVal = rowCell.value
      if not rowVal or type(rowVal) is not str:
        continue
      if 'G/L Week\nEnding' == rowVal:
        headerRow = rowCell.row
        dateCol = rowCell.column
      elif '\nName' == rowVal:
        nameCol = rowCell.column
      elif 'Bill\nRate' == rowVal:
        rateCol = rowCell.column
      elif '\nHours' == rowVal:
        hourCol = rowCell.column

  if headerRow == 0 or dateCol == 0 or nameCol == 0 or rateCol == 0 or hourCol == 0:
    print("Unable to find headers in Deltec xlsx file")
    return(inData)

  row = headerRow + 1
  while ws.cell(row,1).value: 
    if type(ws.cell(row,1).value) is str:
      firstCellVal = ws.cell(row,1).value
      if "Total" in firstCellVal or "TOTAL" in firstCellVal or 'Belcan' in firstCellVal:
        row = row + 1
        continue
    dateVal = ws.cell(row,dateCol).value
    weekNum = dateVal.isocalendar()[1]
    valSplit = ws.cell(row,nameCol).value.split(', ')
    nameSplit = valSplit[1].split(' ')
    nameVal = valSplit[0] + ', ' + nameSplit[0]
    fRateVal = float(ws.cell(row,rateCol).value)
    fHourVal = float(ws.cell(row,hourCol).value)
    if nameVal in inData:
      if weekNum in inData[nameVal]:
        inData[nameVal][weekNum] = float(inData[nameVal][weekNum]) + fHourVal
      else:
        inData[nameVal][weekNum] = fHourVal
      inData[nameVal]['Deltek Rate'] = fRateVal
    else:
      stats = {}
      stats['SpringAhead Rate'] = fRateVal
      stats['Deltek Rate'] = fRateVal
      stats['Avg Run'] = float(40)
      stats["Hours"] = ""
      stats["Dollars"] = ""
      stats[weekNum] = fHourVal
      inData[nameVal] = stats
    row = row + 1

  return(inData)

def readEmployeeFile():
  inData = {}  
  if len(employeeFile) <= 0:
    return(inData)
  with open(employeeFile, newline='') as fileIn:
    inDict = csv.DictReader(fileIn, delimiter=',', quotechar='"')
    for row in inDict:
      usr = row['Name']
      inData[usr] = row

  return(inData)

def readEmployeeExcelFile():
  inData = {}  
  if len(employeeFile) <= 0:
    return(inData)
  excelIn = load_workbook(employeeFile)
  ws = excelIn.worksheets[0]

  nNameCol = 1
  nFirstWeekCol = 2
  nLastWeekCol = 54
  nRow = 1
  while ws.cell(nRow,nNameCol).value: 
    if type(ws.cell(nRow,nNameCol).value) is str:
      nameColVal = ws.cell(nRow,nNameCol).value
      if "Name" in nameColVal or "Available" in nameColVal or "Hours" in nameColVal:
        nRow = nRow + 1
        continue
      valSplit = ws.cell(nRow,nNameCol).value.split(', ')
      nameSplit = valSplit[1].split(' ')
      name = valSplit[0] + ', ' + nameSplit[0]
      projectedHours = {}
      for cells in ws.iter_cols(nFirstWeekCol, nLastWeekCol, nRow, nRow):
        for cell in cells:
          columnName = ws[cell.column_letter + str(1)].value
          projectedHours[columnName] = cell.value
      inData[name] = projectedHours
    nRow = nRow + 1

  return(inData)

def createHistoricalData():
  outputData = {}

  #Calendar data
  calendarData = readCalendarFile()
  calendarRow = {}
  calendarRow["Name"] = "Available Hours"
  calendarRow["SpringAhead Rate"] = calendarRow["Deltek Rate"] = calendarRow["Avg Run"] = calendarRow["Hours"] = calendarRow["Dollars"] = ""
  for w in range(1,54):
    calendarRow[w] = calendarData[w]
  outputData[calendarRow['Name']] = calendarRow

  #SpringAhead historicalData
  springAheadData = readSpringAheadFile()
  global nSpringAheadLastWeekNum
  for name, stats in springAheadData.items():
    outRow = {}
    outRow['Name'] = name
    outRow['SpringAhead Rate'] = stats['SpringAhead Rate']
    outRow['Deltek Rate'] = ""
    outRow['Avg Run'] = stats['Avg Run']
    outRow["Hours"] = ""
    outRow["Dollars"] = ""
    for w in range(1,54):
      if w in stats:
        outRow[w] = float(stats[w])
        if w > nSpringAheadLastWeekNum:
          nSpringAheadLastWeekNum = w
      else:
        outRow[w] = float(0)
    outputData[name] = outRow
  print("SpringAhead Last Week " + str(nSpringAheadLastWeekNum) + " [" + date.fromisocalendar(2020,nSpringAheadLastWeekNum,7).strftime("%m/%d/%y") + "]")

  #Deltek historicalData 
  deltekData = readDeltekExcelFile()
  deltekLastWeek = nSpringAheadLastWeekNum + 1
  for name, stats in deltekData.items():
    outRow = {}
    if name in outputData:
      outRow = outputData[name]
      outRow['Deltek Rate'] = stats['Deltek Rate']
    else:
      outRow['Name'] = name
      outRow['SpringAhead Rate'] = stats['SpringAhead Rate']
      outRow['Deltek Rate'] = stats['Deltek Rate']
      outRow['Avg Run'] = stats['Avg Run']
      outRow["Hours"] = ""
      outRow["Dollars"] = ""
      for w in range(1, 54):
        outRow[w] = float(0)
      outputData[name] = outRow
    for w in range(nSpringAheadLastWeekNum+1,54):
      if w in stats:
        outRow[w] = float(stats[w])
        if w > deltekLastWeek and outRow[w] > float(0):
          deltekLastWeek = w
      else:
        outRow[w] = float(0)
  print("Deltek Last Week " + str(deltekLastWeek) + " [" + date.fromisocalendar(2020,deltekLastWeek,7).strftime("%m/%d/%y") + "]")

  if doProjection == 1:
    # Add any names from employee file that are not in SpringAhead or Deltek files
    empData = readEmployeeExcelFile()
    if empData:
      for name in empData:
        if name not in outputData:
          outRow = {}
          outRow['Name'] = name
          outRow['SpringAhead Rate'] = float(216)
          outRow['Deltek Rate'] = float(216)
          outRow['Avg Run'] = float(40)
          outRow["Hours"] = ""
          outRow["Dollars"] = ""
          for w in range(1, 54):
            outRow[w] = float(0)
          outputData[name] = outRow

  global firstProjectedWeek
  firstProjectedWeek = deltekLastWeek + 1

  return(outputData)

def  radhaFix(outputData):
  outputData['Sahoo, Radha'][32] = float(43)
  outputData['Sahoo, Radha'][33] = float(43)
  outputData['Sahoo, Radha'][34] = float(43)
  outputData['Sahoo, Radha'][35] = float(43)
  outputData['Sahoo, Radha'][36] = float(40)

def setExcelFormulas(ws, rowCnt, colCnt):
  nameCol = 'A'
  springAheadRateCol = 'B'
  deltekRateCol = 'C'
  runRateCol = 'D'
  hourCol = 'E'
  dollarCol = 'F'
  week1Col = 'G'
  week53Col = 'BG'
  sAvailHoursRow = '2'
  nFirstNameRowNum = 3
  sFirstNameRowNum = str(nFirstNameRowNum)
  nLastNameRowNum = rowCnt
  sLastNameRowNum = str(nLastNameRowNum)
  nWeek1ColNum = 7
  nWeek53ColNum = colCnt
  nHourTotRowNum = rowCnt + 1
  sHourTotRowNum = str( nHourTotRowNum)
  nDollarTotRowNum = rowCnt + 2
  sDollarTotRowNum = str(nDollarTotRowNum)
  nLastBilledColumn = firstProjectedWeek + nWeek1ColNum - 2
  sLastBilledColumn = week1Col
  rateCol = springAheadRateCol

  # Total Hours and dollars for each person for the year
  for cells in ws.iter_rows(nFirstNameRowNum,nLastNameRowNum,nWeek1ColNum,nWeek53ColNum):
    sRowNum = str(cells[0].row)
    # Sum Total hours
    formula = '=SUM(' + week1Col + sRowNum  + ':' + week53Col + sRowNum + ')'
    ws[hourCol + sRowNum].value = formula
    # Calculate total dollars from rate and hours per week
    sSpringAheadRateCell = springAheadRateCol + sRowNum
    sDeltekRateCell = deltekRateCol + sRowNum
    formula = '=SUM('
    for cell in cells:
      if cell.column >= (nSpringAheadLastWeekNum + nWeek1ColNum):
        formula += sDeltekRateCell + '*' + cell.column_letter + sRowNum + ','
      else:
        formula += sSpringAheadRateCell + '*' + cell.column_letter + sRowNum + ','
    ws[dollarCol + sRowNum].value = formula[0:len(formula)-1] + ')'

  # Total Hours and Dollars for each week for all people combined
  ws[nameCol + sHourTotRowNum] = 'Hours Total/Week'
  ws[nameCol + sDollarTotRowNum] = 'Dollars Total/Week'
  for cells in ws.iter_cols(nWeek1ColNum, nWeek53ColNum, nFirstNameRowNum, nLastNameRowNum):
    hoursFormula = '=SUM(' + cells[0].coordinate  + ':' + cells[len(cells)-1].coordinate + ')'
    col = cells[0].column_letter
    ws[col + sHourTotRowNum] = hoursFormula
    ws[col + sHourTotRowNum].number_format = '#,##0.00'

    dollarTotalFormula = '=SUM('
    for cell in cells:
      dollarFormula = rateCol + str(cell.row) + '*' + cell.coordinate + ','
      dollarTotalFormula += dollarFormula
    dollarTotalFormula = dollarTotalFormula[0:len(dollarTotalFormula)-1] + ')'
    ws[col + sDollarTotRowNum] = dollarTotalFormula
    ws[col + sDollarTotRowNum].number_format = '"$"#,##0.00'

  # Run Rate for each person for billed weeks  
  nLastBilledColumn = firstProjectedWeek + nWeek1ColNum - 2
  for cells in ws.iter_rows(nFirstNameRowNum,nLastNameRowNum,nWeek1ColNum,nLastBilledColumn):
    numerator = 'SUM('
    denominator = 'SUM('
    hrsTotal = float(0)
    for cell in cells:
      numerator += 'IF(' + cell.coordinate + '>0,' + cell.coordinate + '/$' + cell.column_letter + '$' + sAvailHoursRow + ',0),'
      denominator += 'IF(' + cell.coordinate + '>0,1,0),'
      sLastBilledColumn = cell.column_letter
      hrsTotal += (cell.value)
    numerator = numerator[0:len(numerator)-1] + ')'
    denominator = denominator[0:len(denominator)-1] + ')'
    if hrsTotal > 0:
      ws[runRateCol + str(cells[0].row)] = '=' + numerator + '/' + denominator
    else:
      ws[runRateCol + str(cells[0].row)] = float(1.0)

  # Total hours and dollars for the year all people combined
  nRow = nDollarTotRowNum + 1
  sBudgetRow = str(nRow)
  ws[nameCol + sBudgetRow] = 'Budget'
  ws[dollarCol + sBudgetRow] = 8958790
  ws[dollarCol + sBudgetRow].number_format = '"$"#,##0.00'
  nRow += 1
  sProjYearTotRow = str(nRow)
  ws[nameCol + sProjYearTotRow] = 'Projected Year Totals'
  formula = '=AVERAGE(' + springAheadRateCol + sFirstNameRowNum  + ':' + springAheadRateCol + sLastNameRowNum + ')'
  ws[springAheadRateCol + sProjYearTotRow] = formula
  ws[springAheadRateCol + sProjYearTotRow].number_format = '"$"#,##0.00'
  formula = '=AVERAGE(' + deltekRateCol + sFirstNameRowNum  + ':' + deltekRateCol + sLastNameRowNum + ')'
  ws[deltekRateCol + sProjYearTotRow] = formula
  ws[deltekRateCol + sProjYearTotRow].number_format = '"$"#,##0.00'
  formula = '=AVERAGE(' + runRateCol + sFirstNameRowNum  + ':' + runRateCol + sLastNameRowNum + ')'
  ws[runRateCol + sProjYearTotRow] = formula
  ws[runRateCol + sProjYearTotRow].number_format = '#,##0.00'
  formula = '=SUM(' + hourCol + sFirstNameRowNum  + ':' + hourCol + sLastNameRowNum + ')'
  ws[hourCol + sProjYearTotRow] = formula
  ws[hourCol + sProjYearTotRow].number_format = '#,##0.00'
  formula = '=SUM(' + dollarCol + sFirstNameRowNum  + ':' + dollarCol + sLastNameRowNum + ')'
  ws[dollarCol + sProjYearTotRow] = formula
  ws[dollarCol + sProjYearTotRow].number_format = '"$"#,##0.00'
  nRow += 1
  sYearToDateTotRow = str(nRow)
  formula = '=SUM(' + week1Col + sDollarTotRowNum  + ':' + sLastBilledColumn + sDollarTotRowNum + ')'
  ws[nameCol + sYearToDateTotRow] = 'Year To Date '
  ws[dollarCol + sYearToDateTotRow] = formula
  ws[dollarCol + sYearToDateTotRow].number_format = '"$"#,##0.00'
  nRow += 1
  sETC_Row = str(nRow)
  ws[nameCol + sETC_Row] = 'Estimate To Complete (ETC)'
  ws[dollarCol + sETC_Row] = '=' + dollarCol + sBudgetRow + '-' + dollarCol + sYearToDateTotRow
  ws[dollarCol + sETC_Row].number_format = '"$"#,##0.00'
  nRow += 1
  sEAC_Row = str(nRow)
  ws[nameCol + sEAC_Row] = 'Estimate At Complete (EAC)'
  ws[dollarCol + sEAC_Row] = '=' + dollarCol + sBudgetRow + '-' + dollarCol + sProjYearTotRow
  ws[dollarCol + sEAC_Row].number_format = '"$"#,##0.00'
      
  if doProjection == 1:
    # Project Future hours
    empData = readEmployeeExcelFile()
    for cells in ws.iter_rows(nFirstNameRowNum, nLastNameRowNum, nLastBilledColumn+1, nWeek1ColNum + 52):
      nRow = cells[0].row
      sRow = str(nRow)
      name = ws[nameCol + sRow].value
      for cell in cells:
        availHrsCell = str(cell.column_letter + '$' + sAvailHoursRow)
        runRateCell = str(runRateCol + sRow)
        w = cell.column - nWeek1ColNum + 1
        columnName = "Week " + str(w) + "\n[" + date.fromisocalendar(2020,w,7).strftime("%m/%d/%y") + "]"
        sProjHrs = '40'
        if empData:
          sProjHrs = str(empData[name][columnName])
        normalizedHrs = '(' + sProjHrs + ' - (40 - $' + availHrsCell + '))'
        formula = '=IF(' + normalizedHrs + ' <= 0, 0,' + normalizedHrs + ' * ' + runRateCell + ')'
        ws[cell.coordinate] = formula

def createOutputExcel(outputData):

  wb = Workbook()
  ws = wb.active
  ws.title = 'BudgetProjection'

  columnNames = ["Name","SA Rate","Deltek Rate","Avg Run","Hours","Dollars"]
  for w in range(1,54):
    columnNames.append("Week " + str(w) + " \n[" + date.fromisocalendar(2020,w,7).strftime("%m/%d/%y") + "]") 
  ws.freeze_panes = 'G3'

  headerCnt = len(columnNames)
  for cells in ws.iter_rows(1,1,1,headerCnt):
    for cell in cells:
      cell.value = columnNames[cell.column-1]
      cell.alignment = Alignment(wrapText = True,horizontal='center')

  rowNum = 2
  rowCnt = len(outputData) + 1
  for name, stats in outputData.items():
    for cells in ws.iter_rows(rowNum,rowNum,1,headerCnt):
      cells[0].value = name
      ws.column_dimensions[cells[0].column_letter].width = 25
      cells[1].value = stats['SpringAhead Rate']
      cells[1].number_format = '"$"#'
      ws.column_dimensions[cells[1].column_letter].width = 8
      cells[2].value = stats['Deltek Rate']
      cells[2].number_format = '"$"#'
      ws.column_dimensions[cells[2].column_letter].width = 8
      cells[3].value = stats['Avg Run']
      cells[3].number_format = '0.00'
      ws.column_dimensions[cells[3].column_letter].width = 7
      cells[4].value = stats['Hours']
      cells[4].number_format = '#,##0.00'
      ws.column_dimensions[cells[4].column_letter].width = 9 
      cells[5].value = stats['Dollars']
      cells[5].number_format = '"$"#,##0.00'
      ws.column_dimensions[cells[5].column_letter].width = 14 
      for w in range(1,54):
        cells[w+5].value = float(stats[w])
        if rowNum > rowCnt:
          cells[w+5].number_format = '"$"#,##0.00'
        else:
          cells[w+5].number_format = '#,##0.00'
        ws.column_dimensions[cells[w+5].column_letter].width = 12 
        ws.column_dimensions[cells[w+5].column_letter].height = 29 

    rowNum = rowNum + 1
  
  setExcelFormulas(ws, rowCnt, headerCnt)

  wb.save(outputFile)

if __name__ == '__main__':
  rv = getCmdlineArgs(sys.argv[1:])
  if rv == 1:
    sys.exit(usage)
  elif rv == 2:
    sys.exit("Command line errors\n" + usage)

  print("Calendar file is: '" + calendarFile + "'")
  print("SpringAhead file is: '" + springAheadFile + "'")
  print("Deltek file is: '" + deltekFile + "'")
  print("Output file is: '" + outputFile + "'")

  outputData = createHistoricalData()
  radhaFix(outputData)
  createOutputExcel(outputData)

